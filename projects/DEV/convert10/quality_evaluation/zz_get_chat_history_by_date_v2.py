import os
import json
import boto3
from botocore.exceptions import ClientError
import datetime
import logging

from boto3.dynamodb.conditions import Attr
from decimal import Decimal
import sys
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment
from pathlib import Path

# Dynamically add the parent directory to sys.path
script_dir = Path(__file__).resolve().parent
project_root = script_dir.parents[0]

if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

# Import and load environment variables using the helper function
from utils import load_env
load_env()

# Ensure BASE_DIR is correctly set from the environment
BASE_DIR = Path(os.getenv('BASE_DIR'))

# Define the relative path for the chat history directory
CHAT_HISTORY_DIR = BASE_DIR / 'quality_evaluation/chat_history'

# Ensure the chat history directory exists
CHAT_HISTORY_DIR.mkdir(parents=True, exist_ok=True)

# Custom JSON encoder for handling Decimal types from DynamoDB
class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)  # Convert Decimal to float for JSON serialization
        return super(DecimalEncoder, self).default(obj)

# Configure logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Configuration values from the .env file
DYNAMODB_TABLE_NAME = os.getenv('DYNAMODB_TABLE_NAME')
JSON_FILE_NAME = 'chat_history.json'
PATH_DYNAMODB_CHAT_HISTORY = BASE_DIR / JSON_FILE_NAME
REGION_NAME = os.getenv('AWS_REGION', 'eu-central-1')
START_TIME = os.getenv('START_TIME')  
END_TIME = os.getenv('END_TIME')      

# Function to create a DynamoDB client
def create_dynamodb_client():
    try:
        dynamodb = boto3.resource('dynamodb', region_name=REGION_NAME)
        logger.info("DynamoDB Client successfully created.")
        return dynamodb
    except Exception as e:
        logger.error(f"Error creating DynamoDB client: {e}")
        sys.exit(1)

# Function to convert date strings to UNIX timestamps (in milliseconds)
def to_unix_time_millis(date_str):
    return int(datetime.datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S').timestamp() * 1000)

# Function to scan the DynamoDB table within a specific time range
def scan_table_for_time_range(dynamodb_client, table_name, start_time, end_time):
    try:
        logger.debug(f"Scanning table {table_name} from {start_time} to {end_time}")
        table = dynamodb_client.Table(table_name)
        start_time_millis = to_unix_time_millis(start_time)
        end_time_millis = to_unix_time_millis(end_time)

        response = table.scan(
            FilterExpression=Attr('msg_epoch_time').between(start_time_millis, end_time_millis)
        )
        items = response['Items']
        logger.debug(f"Initial scan returned {len(items)} items")

        while 'LastEvaluatedKey' in response:
            response = table.scan(
                FilterExpression=Attr('msg_epoch_time').between(start_time_millis, end_time_millis),
                ExclusiveStartKey=response['LastEvaluatedKey']
            )
            items.extend(response['Items'])
            logger.debug(f"Scan returned {len(response['Items'])} more items")

        # Add a readable date to each item
        for item in items:
            epoch_time = item.get('msg_epoch_time')
            epoch_time_int = int(float(epoch_time))
            readable_date = datetime.datetime.fromtimestamp(epoch_time_int / 1000).strftime('%Y-%m-%d %H:%M:%S')
            item['readable_date'] = readable_date

        logger.info(f"Records successfully retrieved for the time range from {start_time} to {end_time}. Total items: {len(items)}")
        return items
    except ClientError as e:
        logger.error(f"Error during scan operation: {e}")
        sys.exit(1)

# Function to save records as JSON
def save_records_as_json(records, file_path):
    try:
        with open(file_path, 'w') as file:
            json.dump(records, file, indent=4, cls=DecimalEncoder)
        logger.info(f"Records successfully saved as JSON: {file_path}")
    except Exception as e:
        logger.error(f"Error saving JSON file: {e}")
        sys.exit(1)

# Function to load JSON data
def load_json_data(file_path):
    with open(file_path, 'r') as file:
        return json.load(file)

# Function to check if a message should be excluded
def message_excluded(human_msg, ai_msg, chat_id, exclude_terms):
    exclude_ai_message = "Hallo, ich bin dein persönlicher Assistent. Wie kann ich dir helfen?"
    if any(exclude_term.lower() in human_msg.lower() for exclude_term in exclude_terms):
        return False
    if ai_msg == exclude_ai_message:
        return False
    if "CHAT_ID_RAGAS_" in chat_id:
        return False
    if human_msg.lower().endswith("eoq"):
        return False
    return False

# Function to save messages to an Excel file
def save_to_excel(messages, file_path, exclude_terms):
    logger.debug(f"Saving {len(messages)} messages to Excel")
    wb = Workbook()
    ws = wb.active
    ws.append([
        "readable_date", "HumanMessage", "template", "AIMessage", "enrichment" ,"trace_id" , "chat_id",
        "user_id" , "HumanMessage_token_num", "message_id","messages_token_size", "rate","msg_epoch_time" ,"AIMessage_token_num" , "used_messages_of_all", "vector_results"
        ])

    # Sort messages by msg_epoch_time
    sorted_messages = sorted(messages, key=lambda x: int(x['msg_epoch_time']['N']) if isinstance(x['msg_epoch_time'], dict) else int(x['msg_epoch_time']))

    for message in sorted_messages:
        # Extract fields from message
        user_id = message.get('user_id', {}).get('S', '') if isinstance(message.get('user_id'), dict) else message.get('user_id', '')
        msg_epoch_time = message.get('msg_epoch_time', {}).get('N', '') if isinstance(message.get('msg_epoch_time'), dict) else message.get('msg_epoch_time', '')
        
        # Convert epoch time to readable date
        if msg_epoch_time:
            epoch_time_int = int(float(msg_epoch_time))
            readable_date = datetime.datetime.fromtimestamp(epoch_time_int / 1000).strftime('%Y-%m-%d %H:%M:%S')
        else:
            readable_date = ''
        
        AIMessage = message.get('AIMessage', {}).get('S', '') if isinstance(message.get('AIMessage'), dict) else message.get('AIMessage', '')
        AIMessage_token_num = message.get('AIMessage_token_num', {}).get('N', '') if isinstance(message.get('AIMessage_token_num'), dict) else message.get('AIMessage_token_num', '')
        chat_id = message.get('chat_id', {}).get('S', '') if isinstance(message.get('chat_id'), dict) else message.get('chat_id', '')
        enrichment = message.get('enrichment', {}).get('S', '') if isinstance(message.get('enrichment'), dict) else message.get('enrichment', '')
        HumanMessage = message.get('HumanMessage', {}).get('S', '') if isinstance(message.get('HumanMessage'), dict) else message.get('HumanMessage', '')
        HumanMessage_token_num = message.get('HumanMessage_token_num', {}).get('N', '') if isinstance(message.get('HumanMessage_token_num'), dict) else message.get('HumanMessage_token_num', '')
        message_id = message.get('message_id', {}).get('S', '') if isinstance(message.get('message_id'), dict) else message.get('message_id', '')
        messages_token_size = message.get('messages_token_size', {}).get('N', '') if isinstance(message.get('messages_token_size'), dict) else message.get('messages_token_size', '')
        rate = message.get('rate', {}).get('S', '') if isinstance(message.get('rate'), dict) else message.get('rate', '')
        template = message.get('template', {}).get('S', '') if isinstance(message.get('template'), dict) else message.get('template', '')
        trace_id = message.get('trace_id', {}).get('S', '') if isinstance(message.get('trace_id'), dict) else message.get('trace_id', '')
        used_messages_of_all = message.get('used_messages_of_all', {}).get('S', '') if isinstance(message.get('used_messages_of_all'), dict) else message.get('used_messages_of_all', '')
        vector_results = str(message.get('vector_results', {}).get('L', [])) if isinstance(message.get('vector_results'), dict) else str(message.get('vector_results', []))

        # Append row to worksheet
        ws.append([
            readable_date, HumanMessage, template, AIMessage, enrichment, trace_id , chat_id,
            user_id , HumanMessage_token_num, message_id,messages_token_size, rate,msg_epoch_time ,AIMessage_token_num , used_messages_of_all, vector_results
        ])

    for row in ws.iter_rows(min_row=2, max_col=16, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    wb.save(file_path)
    logger.info(f"Data successfully saved to Excel: {file_path}")

# Main function to process DynamoDB data and save it to Excel
def process_dynamodb_data():
    dynamodb_client = create_dynamodb_client()
    items = scan_table_for_time_range(dynamodb_client, DYNAMODB_TABLE_NAME, START_TIME, END_TIME)
    
    # Get the current date and time
    now = datetime.datetime.now()
    formatted_now = now.strftime('%Y_%m_%d_%H_%M_%S')  # Include hour, minute, and second
    
    # Define the output paths for JSON and Excel files
    output_excel_path = CHAT_HISTORY_DIR / f'chat_history_{formatted_now}.xlsx'
    output_json_path = CHAT_HISTORY_DIR / f'chat_history_{formatted_now}.json'
    
    # Save records as JSON in the same directory as the Excel file
    save_records_as_json(items, output_json_path)
    
    exclude_terms = []
    json_data = load_json_data(output_json_path)
    save_to_excel(json_data, output_excel_path, exclude_terms)

# Program entry point
if __name__ == "__main__":
    process_dynamodb_data()