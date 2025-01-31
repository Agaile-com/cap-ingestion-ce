#!/usr/bin/env python
#02_ragas_evaluation_v0.1.py
# Import required libraries and modules
import csv
import boto3
from botocore.exceptions import ClientError
from tqdm import tqdm
import openpyxl
import re
from datasets import Dataset
from openpyxl import Workbook
import requests
import random
from shutil import copy2
import logging
import pandas as pd
from ragas import evaluate
from ragas.metrics import (
    faithfulness,
    answer_relevancy,
    answer_similarity  # replaces answer correctness
)
import sys
from fuzzywuzzy import fuzz
import datetime
import os
import json
from colorama import init, Fore
from boto3.dynamodb.conditions import Key
from decimal import Decimal
import unicodedata
from dotenv import load_dotenv
import dspy
from pathlib import Path
import numpy as np

# Dynamically add the parent directory to sys.path
script_dir = Path(__file__).resolve().parent
parent_dir = script_dir.parent.parent
sys.path.append(str(parent_dir))

from utils import load_env
load_env()

# Create data directory
data_dir = os.path.join(script_dir, 'data')
os.makedirs(data_dir, exist_ok=True)


# Load necessary variables

OPENAI_API_KEY = os.getenv('OPENAI_API_KEY')
USE_EXCEL = os.getenv('USE_EXCEL', 'true').lower() == 'true'

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Initialize Colorama for colored output
init()

# Normalization function
def normalize_unicode_characters(data):
    if isinstance(data, str):
        return unicodedata.normalize('NFKC', data)
    elif isinstance(data, list):
        return [normalize_unicode_characters(item) for item in data]
    elif isinstance(data, dict):
        return {key: normalize_unicode_characters(value) for key, value in data.items()}
    else:
        return data

# Custom JSON encoder for handling Decimal types from DynamoDB
class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)  # Convert Decimal to float for JSON serialization
        return super(DecimalEncoder, self).default(obj)

# Function für colored output
def colored_print(color, message):
    print(color + message + Fore.RESET)

# Configure logging system
logger = logging.getLogger()
logger.setLevel(logging.DEBUG)

# Create a file handler that logs DEBUG and higher levels
file_handler = logging.FileHandler('data_evaluation.log')
file_handler.setLevel(logging.DEBUG)
file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(file_formatter)
logger.addHandler(file_handler)

# Create a stream handler (console) that logs INFO and higher levels
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
console_handler.setFormatter(console_formatter)
logger.addHandler(console_handler)

# Load configuration paths from environment variables
config_paths = {
    'X_API_KEY': os.getenv('X_API_KEY'),
    'DYNAMODB_TABLE_NAME': os.getenv('DYNAMODB_TABLE_NAME'),
    'CLOUDFRONT_URL': os.getenv('CLOUDFRONT_URL'),
    'OPENAI_API_KEY': os.getenv('OPENAI_API_KEY'),
    'USE_EXCEL': os.getenv('USE_EXCEL', 'true') == 'true',
    'ASSISTANT': os.getenv('ASSISTANT'),
    'SAMPLE_SIZE': int(os.getenv('SAMPLE_SIZE', '0')),
    'PATH_TEST_DATA_QUESTIONS': os.path.join(data_dir, "1_test_data_set_questions_ground_truths.xlsx" if os.getenv('USE_EXCEL', 'true') == 'true' else "1_test_data_set_questions_ground_truths.csv"),
    'PATH_TEST_DATA_RAG_ANSWERS': os.path.join(data_dir, "2_test_data_set_rag_answers.xlsx" if os.getenv('USE_EXCEL', 'true') == 'true' else "2_test_data_set_rag_answers.csv"),
    'PATH_DYNAMODB_CHAT_HISTORY': os.path.join(data_dir, 'dynamoDB_test_set_chat_history.json'),
    'PATH_RESULTS_EVAL_RAGAS': os.path.join(data_dir, "results_eval_ragas.xlsx" if os.getenv('USE_EXCEL', 'true') == 'true' else "results_eval_ragas.csv"),
}

colored_print(Fore.CYAN, "Loaded configuration paths from .env file:")
for key, value in config_paths.items():
    colored_print(Fore.YELLOW, f"{key}: {value}")

# Dynamische Modellkonfiguration basierend auf der Umgebungsvariable
dspy.configure(lm=dspy.OpenAI(model='gpt-4', api_key=config_paths['OPENAI_API_KEY']))

# Klassenbasierte DSPy-Signatur für die Fragegenerierung
class GenerateQuestion(dspy.Signature):
    """
    Your task is to formulate a concise and direct chat request. Use simple, universally understandable words. These requests should be brief and focused on information gathering. Base the chat request on the content summary rather than specific details. Formulate only one chat request at a time.
    """
    content = dspy.InputField(desc="Base for chat request")
    question = dspy.OutputField(desc="Chatbot request")

generate_question_model = dspy.ChainOfThought(GenerateQuestion)

def generate_question(content):
    response = generate_question_model(content=content)
    return response.question.strip()

# Function to create a DynamoDB client
def create_dynamodb_client():
    try:
        dynamodb = boto3.resource('dynamodb', region_name='eu-central-1')
        logging.info("DynamoDB Client successfully created.")
        return dynamodb
    except Exception as e:
        logging.error(f"Error creating DynamoDB Client: {e}")
        sys.exit(1)

# Function to scan, sort, and select the n most recent records from a DynamoDB table
def scan_and_sort_table(dynamodb_client, table_name, user_id):
    try:
        table = dynamodb_client.Table(table_name)
        response = table.query(
            KeyConditionExpression=Key("user_id").eq(user_id),
        )
        items = response['Items']

        while 'LastEvaluatedKey' in response:
            response = table.query(
                KeyConditionExpression=boto3.dynamodb.conditions.Key("user_id").eq(user_id),
                ExclusiveStartKey=response['LastEvaluatedKey']
            )
            items.extend(response['Items'])

        for item in items:
            epoch_time = item.get('msg_epoch_time')
            epoch_time_ms = int(epoch_time) / 1000 if len(str(epoch_time)) == 13 else int(epoch_time)
            readable_date = datetime.datetime.fromtimestamp(epoch_time_ms).strftime('%Y-%m-%d %H:%M:%S')
            item['readable_date'] = readable_date

        sorted_items = sorted(items, key=lambda x: x['msg_epoch_time'], reverse=True)

        logging.info(f"The {len(items)} most recent records were successfully retrieved and sorted.")
        return sorted_items
    except ClientError as e:
        logging.error(f"Error during scan operation: {e}")
        sys.exit(1)

# Function to save the n most recent records as JSON
def save_n_records_as_json(records, file_path):
    normalized_records = normalize_unicode_characters(records)
    try:
        with open(file_path, 'w', encoding='utf-8') as file:
            json.dump(normalized_records, file, indent=4, ensure_ascii=False, cls=DecimalEncoder)
        logging.info(f"The normalized most recent records were successfully saved as JSON: {file_path}")
    except Exception as e:
        logging.error(f"Error saving the normalized JSON file: {e}")
        sys.exit(1)

# Main function to process DynamoDB data
def process_dynamodb_data(user_id):
    dynamodb_client = create_dynamodb_client()
    table_name = config_paths['DYNAMODB_TABLE_NAME']
    logging.info(f"Using DynamoDB table: {table_name}")
    sorted_data = scan_and_sort_table(dynamodb_client, table_name, user_id)
    json_file_path = config_paths['PATH_DYNAMODB_CHAT_HISTORY']
    save_n_records_as_json(sorted_data, json_file_path)

# Function to copy and rename a file
def copy_and_rename_file(source_path, destination_path):
    try:
        logging.debug(f"Starting copy_and_rename_file from {source_path} to {destination_path}")
        source_wb = openpyxl.load_workbook(source_path)
        source_sheet = source_wb.active
        logging.debug(f"Source Excel structure: Rows={source_sheet.max_row}, Columns={source_sheet.max_column}")
        
        # Debug jede Zeile der Quell-Excel
        for idx, row in enumerate(source_sheet.rows, 1):
            logging.debug(f"Source Excel Row {idx}: {[cell.value for cell in row]}")
        
        dest_wb = openpyxl.Workbook()
        dest_sheet = dest_wb.active
        
        # Kopiere Zelle für Zelle
        for row in source_sheet.rows:
            for cell in row:
                dest_sheet[cell.coordinate] = cell.value
                logging.debug(f"Copying cell {cell.coordinate}: '{cell.value}' to destination")
                
        # Speichere das neue Workbook
        dest_wb.save(destination_path)
        
        # Verify immediately after saving
        verify_wb = openpyxl.load_workbook(destination_path, data_only=True)
        verify_sheet = verify_wb.active
        logging.debug(f"Immediate verification after save:")
        logging.debug(f"Cell A2 value: '{verify_sheet['A2'].value}'")
        logging.debug(f"Cell A2 data type: {type(verify_sheet['A2'].value)}")
        logging.debug(f"Cell A2 number format: {verify_sheet['A2'].number_format}")
        logging.debug(f"Cell A2 style properties: {verify_sheet['A2']._style}")
        
        return True
    except Exception as e:
        logging.error(f"Error in copy_and_rename_file: {e}", exc_info=True)
        return False

# Function to read data from a file
def read_data(file_path, use_excel=True, samplesize='all'):
    try:
        logging.debug(f"Starting read_data from {file_path}")
        
        if use_excel:
            # Read the Excel file using pandas instead of openpyxl
            df = pd.read_excel(file_path)
            
            # Convert to list of tuples (question, ground_truth)
            data = list(zip(df['question'].tolist(), df['ground_truth'].tolist()))
            
            logging.debug(f"Read {len(data)} rows from Excel")
            logging.debug(f"First row sample: {data[0]}")
            
            if samplesize != 'all':
                samplesize = int(samplesize)
                data = data[:samplesize]
                
            return data
            
    except Exception as e:
        logging.error(f"Error in read_data: {e}")
        raise

def create_new_chat(user_id):
    try:
        chat_id = f"CHAT_ID_RAGAS_{random.randint(1, 9999999)}"
        url = f"{config_paths['CLOUDFRONT_URL']}new_chat"

        headers = {
            "Content-Type": "application/json; charset=utf-8",
            "x-api-key": config_paths['X_API_KEY']
        }
        response = requests.post(url, json={"chatID": chat_id, "userID": user_id}, headers=headers)

        if response is None:
            logging.error("No response received from backend.")
            return None, None
        elif "chatID" not in response.json():
            logging.error(f"ChatID missing in server response. Full response: {response.json()}")
            return None, None
        elif response.json()["chatID"] != chat_id:
            logging.error(f"Invalid ChatID in server response. Received ChatID: {response.json()['chatID']}, Expected ChatID: {chat_id}")
            return None, None
        else:
            logging.info("New chat successfully created.")
            return chat_id, user_id
    except Exception as e:
        logging.error(f"Error making network request for /new_chat: {e}")
        return None, None

# Function to get a response from the backend
def get_answer_from_backend(question, chat_id, user_id):
    try:
        url = f"{config_paths['CLOUDFRONT_URL']}chat_answer"
        headers = {
            "Content-Type": "application/json; charset=utf-8",
            "x-api-key": config_paths['X_API_KEY']
        }
        payload = {
            "chatID": chat_id,
            "userID": user_id,
            #"modelID": config_paths['MODEL_ID'],
            "assistant": config_paths['ASSISTANT'],
            "message_content": question,
            "streaming": False,
            "show_links": True
        }
        response = requests.post(url, json=payload, headers=headers)

        logging.debug(f"Server response received: Statuscode={response.status_code}, Content={response.text[:500]}")

        if response.status_code == 200:
            try:
                response_json = response.json()
                message = response_json.get('message', '')
                links = ','.join(response_json.get('links', ['NA']))
                logging.info("Response from backend successfully received.")
                return message, links
            except json.JSONDecodeError:
                logging.error("Error parsing JSON response. Response text: " + response.text)
                return '', 'NA'
        else:
            logging.error(f"Error receiving response for /chat_answer: HTTP status {response.status_code}, Response text: {response.text}")
            return '', 'NA'
    except Exception as e:
        logging.error(f"Error making network request for /chat_answer: {e}")
        return '', 'NA'

# Function to save results in Excel
def save_results_excel(answers, answer_links, file_path):
    """
    Add new columns to existing Excel file with answers and links.
    """
    try:
        # Read existing Excel file
        existing_df = pd.read_excel(file_path)
        logging.info(f"Read existing Excel with {len(existing_df)} rows")

        # Add new columns
        existing_df['answer'] = answers
        existing_df['contexts'] = [''] * len(existing_df)  # Empty column for later DynamoDB population
        existing_df['answer_links'] = answer_links

        # Save back to Excel
        existing_df.to_excel(file_path, index=False)
        logging.info(f"Results successfully saved to '{file_path}'")
        
    except Exception as e:
        logging.error(f"Error saving results to Excel: {e}")
        logging.error("Data sample:")
        logging.error(f"Number of answers: {len(answers)}")
        logging.error(f"Number of links: {len(answer_links)}")
        raise

# Function to save results in CSV
def save_results_csv(question, ground_truth, answer, target_file, answer_links):
    try:
        fieldnames = ['question', 'answer', 'ground_truth', 'contexts', 'answer_links']
        
        with open(target_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for q, gt, ans, links in zip(question, ground_truth, answer, answer_links):
                writer.writerow({
                    'question': normalize_unicode_characters(q),
                    'answer': normalize_unicode_characters(ans),
                    'ground_truth': normalize_unicode_characters(gt),
                    'contexts': '',  # Will be populated later from DynamoDB
                    'answer_links': normalize_unicode_characters(links)
                })

        logging.info(f"Results successfully saved in '{target_file}'.")
    except Exception as e:
        logging.error(f"Error saving results: {e}")

# Function to clean text
def clean_text(text):
    if isinstance(text, str):
        text = text.replace('\n', ' ').strip()
    return text

# Function to compare similarity between two texts
def similar(a, b):
    return fuzz.ratio(a, b) / 100.0

def split_enrichment_contexts(text):
    contexts = re.split(r'\n\n', text)
    return contexts

# Main function to match and enrich data
def process_and_match_data(input_file, results_file, use_excel=True):
    try:
        # Debug-Log für die Input-Daten
        input_data = read_data(input_file, use_excel)
        logging.debug(f"Input data for processing: {input_data}")
        
        with open(config_paths['PATH_DYNAMODB_CHAT_HISTORY'], 'r', encoding='utf-8') as file:
            json_data = json.load(file)
            logging.debug(f"JSON data loaded: {json_data[:100]}...")  # First 100 chars
            
        df_json = pd.DataFrame(json_data)
        df_json['HumanMessage'] = df_json['HumanMessage'].apply(
            lambda x: normalize_unicode_characters(clean_text(x)) if isinstance(x, str) else None)
        df_json['AIMessage'] = df_json['AIMessage'].apply(
            lambda x: normalize_unicode_characters(clean_text(x)) if isinstance(x, str) else None)
        df_json['enrichment_contexts'] = df_json['enrichment'].apply(
            lambda x: [normalize_unicode_characters(context) for context in split_enrichment_contexts(x)] if isinstance(x, str) else [])

        dtype_dict = {'question': str}
        df = pd.read_excel(input_file, dtype=dtype_dict) if use_excel else pd.read_csv(input_file, dtype=dtype_dict)
        logging.debug(f"DataFrame after reading Excel: {df.head().to_dict()}")
        
        df['answer'] = df['answer'].apply(lambda x: normalize_unicode_characters(clean_text(x)))
        df['ground_truth'] = df['ground_truth'].apply(lambda x: normalize_unicode_characters(clean_text(x)))
        df['question'] = df['question'].apply(lambda x: normalize_unicode_characters(clean_text(x)))
        logging.debug(f"DataFrame after normalization: {df.head().to_dict()}")

        df['question'] = df['question'].astype(str)
        df['answer'] = df['answer'].astype(str)
        df['answer_links'] = df['answer_links'].astype(str)
        logging.debug(f"DataFrame after type conversion: {df.head().to_dict()}")

        df['question'] = df['question'].astype(str)
        df['answer'] = df['answer'].astype(str)
        df['answer_links'] = df['answer_links'].astype(str)

        matched_results = []
        match_count = 0
        matching_threshold = 0.9

        for index, data_row in tqdm(df.iterrows(), total=df.shape[0], desc="Matching progress"):
            data_question = data_row['question']
            best_match = None
            highest_similarity = 0

            for _, json_row in df_json.iterrows():
                json_humanmessage = json_row['HumanMessage']
                similarity_score = similar(data_question, json_humanmessage)
                if similarity_score > highest_similarity:
                    best_match = json_row
                    highest_similarity = similarity_score

            if highest_similarity >= matching_threshold:
                match_count += 1
                data_row['contexts'] = [normalize_unicode_characters(clean_text(context)) for context in best_match['enrichment_contexts']]
                matched_results.append(data_row.to_dict())
            else:
                pass

            tqdm.write(f"Matches found: {match_count}/{index + 1}")

        if match_count == 0:
            logging.warning("No matches found. Script will terminate.")
            sys.exit(0)

        if use_excel:
            final_excel_path = results_file
            wb = Workbook()
            ws = wb.active
            ws.append(['question', 'answer', 'ground_truth', 'contexts', 'answer_links'])  # Fixed column order

            for record in matched_results:
                ws.append([
                    record['question'],
                    record['answer'],  # Changed position
                    record['ground_truth'],
                    '; '.join(record.get('contexts', [])),
                    record.get('answer_links', 'NA') 
                ])
            wb.save(final_excel_path)
        else:
            final_csv_path = results_file
            control_file_path = final_csv_path.replace('.csv', '_control.csv')

            with open(final_csv_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                fieldnames = ['question', 'ground_truth', 'answer', 'contexts','answer_links' ]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

                writer.writeheader()
                for record in matched_results:
                    record = {
                        'question': record['question'],
                        'ground_truth': record['ground_truth'],
                        'answer': record['answer'],
                        'answer_links': record['answer_links'],
                        'contexts': '; '.join(record.get('contexts', []))
                    }
                    try:
                        writer.writerow(record)
                    except Exception as e:
                        logging.error(f"Error saving record to final_csv_path: {e}\n{record}")
                        continue

            with open(control_file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                fieldnames = ['csv_question', 'human_message', 'matched_answer', 'contexts', 'Match Confirmation']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                for record in matched_results:
                    human_message = None
                    best_similarity_score = 0
                    for _, item in df_json.iterrows():
                        current_score = similar(record['question'], item['HumanMessage'])
                        if current_score > best_similarity_score:
                            best_similarity_score = current_score
                            human_message = item['HumanMessage']

                    match_confirmation = "OK"
                    if best_similarity_score < 0.7:
                        match_confirmation = "Check it!"

                    try:
                        writer.writerow({
                            'csv_question': record['question'],
                            'human_message': human_message or "No match found",
                            'matched_answer': record['answer'],
                            'contexts': '; '.join(record.get('contexts', [])),
                            'Match Confirmation': match_confirmation
                        })
                    except Exception as e:
                        logging.error(f"Error saving record to control_file_path: {e}\n{record}")
                        continue
    except Exception as e:
        logging.error(f"Error in process_and_match_data: {e}")
        raise

# Function to evaluate RAGAS results
def evaluate_ragas(input_file, use_excel=True):
    try:
        openai_api_key = config_paths["OPENAI_API_KEY"]
        if openai_api_key is None:
            logging.error("OPENAI_API_KEY not found.")
            sys.exit(1)
        os.environ["OPENAI_API_KEY"] = openai_api_key

        # Read the data
        df = pd.read_excel(input_file) if use_excel else pd.read_csv(input_file)
        
        # Store answer_links before any processing
        answer_links = df['answer_links'].copy()
        
        # Replace deprecated applymap with map
        for column in df.columns:
            if df[column].dtype == 'object':
                df[column] = df[column].map(lambda x: normalize_unicode_characters(x) if isinstance(x, str) else x)
        
        # Create a copy instead of a view
        df_ragas = df.copy()
        
        # Convert types using loc to avoid SettingWithCopyWarning
        df_ragas.loc[:, 'question'] = df_ragas['question'].astype(str)
        df_ragas.loc[:, 'answer'] = df_ragas['answer'].astype(str)
        df_ragas.loc[:, 'ground_truth'] = df_ragas['ground_truth'].astype(str)
        
        # Process contexts
        def process_contexts(x):
            if pd.isna(x):
                return []
            if isinstance(x, str):
                return [s.strip() for s in x.split(';') if s.strip()]
            if isinstance(x, list):
                return [str(item) for item in x]
            return []
            
        df_ragas.loc[:, 'contexts'] = df_ragas['contexts'].apply(process_contexts)

        # Rename columns for Ragas
        df_for_ragas = df_ragas.rename(columns={
            'question': 'user_input',
            'answer': 'response',
            'ground_truth': 'reference',
            'contexts': 'retrieved_contexts'
        })

        # Create dataset and evaluate
        dataset = Dataset.from_pandas(df_for_ragas)
        result = evaluate(
            dataset=dataset,
            metrics=[
                #faithfulness,
                answer_relevancy,
                answer_similarity
            ]
        )

        logging.info("Evaluation completed successfully.")

        # Convert results back to DataFrame
        result_df = result.to_pandas()
        
        # Add back the answer_links
        result_df['answer_links'] = answer_links
        
        # Save with all columns
        if use_excel:
            result_df.to_excel(input_file, index=False)
        else:
            result_df.to_csv(input_file, index=False)
            
        logging.info(f"Saved results with columns: {result_df.columns.tolist()}")
            
    except Exception as e:
        logging.error(f"Error during evaluation: {e}", exc_info=True)
        sys.exit(1)

# ---- START main program ---- #

# Main program
use_excel_files = config_paths['USE_EXCEL']
source_file_path = config_paths['PATH_TEST_DATA_QUESTIONS']
destination_file_path = config_paths['PATH_TEST_DATA_RAG_ANSWERS']

# Copy and rename file
copy_and_rename_file(source_file_path, destination_file_path)

# Read data
data = read_data(destination_file_path, use_excel_files, 'all')

question, ground_truth = zip(*data)
answers = []
answer_links = []

current_run_user_id = f"USER_ID_RAGAS_{random.randint(1, 9999999)}"
logger.info(f"#################### Running RAGAS script for user ID: {current_run_user_id} ####################\n")
for question in tqdm(question, desc='Processing questions'):
    chat_id, current_run_user_id = create_new_chat(current_run_user_id)
    if chat_id:
        answer, links = get_answer_from_backend(question, chat_id, current_run_user_id)
        answers.append(normalize_unicode_characters(answer) if answer is not None else '')
        answer_links.append(normalize_unicode_characters(links))
    else:
        answers.append('')
        answer_links.append('NA')


# Save results
if use_excel_files:
    save_results_excel(answers, answer_links, destination_file_path)
else:
    save_results_csv(question, ground_truth, answers, destination_file_path, answer_links)

# read from DynamoDB
process_dynamodb_data(current_run_user_id)

# Match and enrich data
process_and_match_data(input_file=config_paths['PATH_TEST_DATA_RAG_ANSWERS'],
                       results_file=config_paths['PATH_RESULTS_EVAL_RAGAS'],
                       use_excel=use_excel_files)

# Evaluate RAGAS results
evaluate_ragas(input_file=config_paths['PATH_RESULTS_EVAL_RAGAS'], use_excel=use_excel_files)

# Function to convert CSV to XLSX
def convert_csv_to_xlsx(csv_file_path, xlsx_file_path):
    try:
        logging.info(f'Starting to convert {csv_file_path} to XLSX format.')
        if csv_file_path.endswith('.xlsx'):
            # If input is already XLSX, just copy it
            logging.info(f'File is already in XLSX format.')
            return
        
        csv_data = pd.read_csv(csv_file_path, encoding='utf-8')
        csv_data.to_excel(xlsx_file_path, index=False)
        logging.info(f'Successfully converted {csv_file_path} to {xlsx_file_path}.')
    except Exception as e:
        logging.error(f'Error converting {csv_file_path} to XLSX: {e}')
        sys.exit(1)

final_csv_path = config_paths['PATH_RESULTS_EVAL_RAGAS']
now = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
final_xlsx_path = final_csv_path.replace('.csv', f'_{now}.xlsx')

convert_csv_to_xlsx(final_csv_path, final_xlsx_path)
# ---- END main program ---- #