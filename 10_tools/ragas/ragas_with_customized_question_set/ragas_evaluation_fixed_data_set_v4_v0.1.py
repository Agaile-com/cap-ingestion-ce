#!/usr/bin/env python
#02_ragas_evaluation_v4.py

# Importieren der benötigten Bibliotheken und Module
import csv
import boto3
from botocore.exceptions import ClientError
from tqdm import tqdm
import openpyxl
import re
from datasets import Dataset
from openpyxl import Workbook
import requests
import random
from shutil import copy2
import logging
import pandas as pd
from ragas import evaluate
from ragas.metrics import context_relevancy, answer_relevancy, faithfulness, context_recall
import sys
from fuzzywuzzy import fuzz
import datetime
import os
import json
from colorama import init, Fore
from boto3.dynamodb.conditions import Key
from decimal import Decimal
import unicodedata
from dotenv import load_dotenv
import dspy
from xlsx2csv import Xlsx2csv
from pathlib import Path

# Dynamically add the parent directory to sys.path
script_dir = Path(__file__).resolve().parent
parent_dir = script_dir.parent
sys.path.append(str(parent_dir))

# Load environment variables using the helper function
from utils import load_env
load_env()

# Initialize Colorama for colored output
init()

# Initialize logging
def setup_logging():
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

    file_handler = logging.FileHandler('data_evaluation.log')
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    return logger

logger = setup_logging()

# Define directories
input_data_dir = Path('ragas_input_data')  # Directory for input files
results_dir = Path('ragas_results')        # Directory for output files
results_dir.mkdir(exist_ok=True)

# Get the input file from the input data directory
input_file = next(input_data_dir.glob('*.xlsx'), None)
if not input_file:
    logger.error("No Excel file found in the input data directory.")
    sys.exit(1)

# Generate the output file name
timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
output_file = results_dir / f"results_{timestamp}_{input_file.name}"

def normalize_unicode_characters(data):
    if isinstance(data, str):
        return unicodedata.normalize('NFKC', data)
    elif isinstance(data, list):
        return [normalize_unicode_characters(item) for item in data]
    elif isinstance(data, dict):
        return {key: normalize_unicode_characters(value) for key, value in data.items()}
    else:
        return data

# Custom JSON encoder for handling Decimal types from DynamoDB
class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)  # Convert Decimal to float for JSON serialization
        return super(DecimalEncoder, self).default(obj)

# Funktion für farbige Ausgabe
def colored_print(color, message):
    print(color + message + Fore.RESET)

# Konfiguration des Logging-Systems
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# Erstellt einen File Handler, der DEBUG und höhere Level loggt und in einer Datei speichert
file_handler = logging.FileHandler('data_evaluation.log')
file_handler.setLevel(logging.DEBUG)
file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(file_formatter)
logger.addHandler(file_handler)

# Erstellt einen Stream Handler (Konsole), der INFO und höhere Level loggt
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
console_handler.setFormatter(console_formatter)
logger.addHandler(console_handler)

# Funktion zur Konvertierung von Excel zu CSV
def convert_excel_to_csv(excel_path, csv_path):
    try:
        # Convert Excel file to CSV
        logging.info(f'Reading Excel file from {excel_path}')
        Xlsx2csv(excel_path, outputencoding="utf-8").convert(csv_path)
        logging.info(f'Data saved to CSV file {csv_path}')
        logging.info('Conversion successful.')
    except Exception as e:
        logging.error(f'Error during conversion: {e}')
        sys.exit(1)

# Konvertieren der Excel-Datei in eine CSV-Datei
convert_excel_to_csv(os.getenv('EXCEL_FILE_PATH_FIXED_DATA'), os.getenv('CSV_FILE_PATH_FIXED_DATA'))

# Dynamische Modellkonfiguration basierend auf der Umgebungsvariable
dspy.configure(lm=dspy.OpenAI(model='gpt-4', api_key=os.getenv('OPENAI_API_KEY')))

# Funktion zur Erstellung eines DynamoDB-Clients
def create_dynamodb_client():
    try:
        dynamodb = boto3.resource('dynamodb', region_name=os.getenv('AWS_REGION'))
        logging.info("DynamoDB Client erfolgreich erstellt.")
        return dynamodb
    except Exception as e:
        logging.error(f"Fehler bei der Erstellung des DynamoDB Clients: {e}")
        sys.exit(1)

# Funktion zum Scannen, Sortieren und Auswählen der n neuesten Datensätze einer DynamoDB-Tabelle
def scan_and_sort_table(dynamodb_client, table_name, user_id):
    try:
        table = dynamodb_client.Table(table_name)
        response = table.query(
            KeyConditionExpression=Key("user_id").eq(user_id),
        )
        items = response['Items']

        while 'LastEvaluatedKey' in response:
            response = table.query(
                KeyConditionExpression=boto3.dynamodb.conditions.Key("user_id").eq(user_id),
                ExclusiveStartKey=response['LastEvaluatedKey']
            )
            items.extend(response['Items'])

        for item in items:
            epoch_time = item.get('msg_epoch_time')
            epoch_time_ms = int(epoch_time) / 1000 if len(str(epoch_time)) == 13 else int(epoch_time)
            readable_date = datetime.datetime.fromtimestamp(epoch_time_ms).strftime('%Y-%m-%d %H:%M:%S')
            item['readable_date'] = readable_date

        sorted_items = sorted(items, key=lambda x: x['msg_epoch_time'], reverse=True)

        logging.info(f"Die {len(items)} neuesten Datensätze erfolgreich abgerufen und sortiert.")
        return sorted_items
    except ClientError as e:
        logging.error(f"Fehler bei der Scan-Operation: {e}")
        sys.exit(1)

# Funktion zum Speichern der n neuesten Datensätze als JSON
def save_n_records_as_json(records, file_path):
    normalized_records = normalize_unicode_characters(records)
    try:
        with open(file_path, 'w', encoding='utf-8') as file:
            json.dump(normalized_records, file, indent=4, ensure_ascii=False, cls=DecimalEncoder)
        logging.info(f"Die normalisierten neuesten Datensätze wurden erfolgreich als JSON gespeichert: {file_path}")
    except Exception as e:
        logging.error(f"Fehler beim Speichern der normalisierten JSON-Datei: {e}")
        sys.exit(1)

# Hauptfunktion zur Verarbeitung von DynamoDB-Daten
def process_dynamodb_data(user_id):
    dynamodb_client = create_dynamodb_client()
    table_name = os.getenv('DYNAMODB_TABLE_NAME')
    logging.info(f"Verwendete DynamoDB Tabelle: {table_name}")
    sorted_data = scan_and_sort_table(dynamodb_client, table_name, user_id)
    json_file_path = os.getenv('PATH_DYNAMODB_CHAT_HISTORY')
    save_n_records_as_json(sorted_data, json_file_path)

# Funktion zum Kopieren und Umbenennen einer Datei
def copy_and_rename_file(source_path, destination_path):
    try:
        copy2(source_path, destination_path)
        logging.info(f"Datei kopiert und umbenannt von '{source_path}' zu '{destination_path}'.")
    except Exception as e:
        logging.error(f"Fehler beim Kopieren und Umbenennen der Datei: {e}")

# Funktion zum Einlesen der Daten aus einer Datei
def read_data(file_path, use_excel=True, samplesize='all'):
    try:
        if use_excel:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            data = [(row[0].value, row[1].value) for row in sheet.iter_rows(min_row=2) if row[0].value and row[1].value]
        else:
            with open(file_path, encoding='utf-8') as csvfile:
                csv_reader = csv.DictReader(csvfile)
                data = [(row['question'], row['ground_truths']) for row in csv_reader if row['question'] and row['ground_truths']]

        if samplesize != 'all':
            samplesize = int(samplesize)
            data = data[:samplesize]

        logging.info(f"{len(data)} Datensätze erfolgreich aus der Datei gelesen.")
        return data
    except Exception as e:
        logging.error(f"Fehler beim Lesen der Datei: {e}")
        return []

def create_new_chat(user_id):
    try:
        chat_id = f"CHAT_ID_RAGAS_{random.randint(1, 9999999)}"
        url = f"{os.getenv('API_ENDPOINT_BASE_URL')}new_chat"

        headers = {
            "Content-Type": "application/json; charset=utf-8",
            "x-api-key": os.getenv('X_API_KEY')
        }
        response = requests.post(url, json={"chatID": chat_id, "userID": user_id}, headers=headers)

        if response is None:
            logging.error("Keine Antwort vom Backend erhalten.")
            return None, None
        elif "chatID" not in response.json():
            logging.error(f"ChatID fehlt in der Serverantwort. Vollständige Antwort: {response.json()}")
            return None, None
        elif response.json()["chatID"] != chat_id:
            logging.error(f"Ungültige ChatID in der Serverantwort. Erhaltene ChatID: {response.json()['chatID']}, Erwartete ChatID: {chat_id}")
            return None, None
        else:
            logging.info("Neuer Chat erfolgreich erstellt.")
            return chat_id, user_id
    except Exception as e:
        logging.error(f"Fehler bei der Netzwerkanfrage für /new_chat: {e}")
        return None, None

# Funktion zum Erhalten einer Antwort vom Backend
def get_answer_from_backend(question, chat_id, user_id):
    try:
        url = f"{os.getenv('API_ENDPOINT_BASE_URL')}chat_answer"
        headers = {
            "Content-Type": "application/json; charset=utf-8",
            "x-api-key": os.getenv('X_API_KEY')
        }
        payload = {
            "chatID": chat_id,
            "userID": user_id,
            "modelID": os.getenv('MODEL_ID'),
            "assistant_mode": os.getenv('ASSISTANT_MODE'),
            "message_content": question,
            "streaming": False,
            "show_links": True
        }
        response = requests.post(url, json=payload, headers=headers)

        logging.debug(f"Serverantwort erhalten: Statuscode={response.status_code}, Inhalt={response.text[:500]}")

        if response.status_code == 200:
            try:
                response_json = response.json()
                message = response_json.get('message', '')
                links = ','.join(response_json.get('links', ['NA']))
                logging.info("Antwort vom Backend erfolgreich erhalten.")
                return message, links
            except json.JSONDecodeError:
                logging.error("Fehler beim Parsen der JSON-Antwort. Antworttext: " + response.text)
                return '', 'NA'
        else:
            logging.error(f"Fehler beim Erhalten der Antwort für /chat_answer: HTTP-Status {response.status_code}, Antworttext: {response.text}")
            return '', 'NA'
    except Exception as e:
        logging.error(f"Fehler bei der Netzwerkanfrage für /chat_answer: {e}")
        return '', 'NA'

# Funktion zum Speichern der Ergebnisse in Excel
def save_results_excel(questions, ground_truths, answers, file_path):
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet['A1'] = 'question'
        sheet['B1'] = 'ground_truths'
        sheet['C1'] = 'answer'
        for i, (question, ground_truth, answer) in enumerate(zip(questions, ground_truths, answers), start=2):
            sheet[f'A{i}'] = normalize_unicode_characters(question)
            sheet[f'B{i}'] = normalize_unicode_characters(ground_truth)
            sheet[f'C{i}'] = normalize_unicode_characters(answer)
        workbook.save(file_path)
        logging.info(f"Ergebnisse erfolgreich in '{file_path}' gespeichert.")
    except Exception as e:
        logging.error(f"Fehler beim Speichern der Ergebnisse: {e}")

# Funktion zum Speichern der Ergebnisse in CSV
def save_results_csv(questions, ground_truths, answers, target_file, answer_links):
    try:
        fieldnames = ['question', 'ground_truths', 'answer', 'answer_links']

        with open(target_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()

            for question, ground_truth, answer, links in zip(questions, ground_truths, answers, answer_links):
                writer.writerow({
                    'question': normalize_unicode_characters(question),
                    'ground_truths': normalize_unicode_characters(ground_truth),
                    'answer': normalize_unicode_characters(answer),
                    'answer_links': normalize_unicode_characters(links)
                })

            logging.info(f"Ergebnisse erfolgreich in '{target_file}' gespeichert.")
    except Exception as e:
        logging.error(f"Fehler beim Speichern der Ergebnisse: {e}")

# Funktion zur Textbereinigung
def clean_text(text):
    if isinstance(text, str):
        text = text.replace('\n', ' ').strip()
    return text

# Funktion zum Vergleichen der Ähnlichkeit zweier Texte
def similar(a, b):
    return fuzz.ratio(a, b) / 100.0

def split_enrichment_contexts(text):
    contexts = re.split(r'\n\n', text)
    return contexts

# Hauptfunktion zum Abgleichen und Anreichern von Daten
def process_and_match_data(input_file, results_file, use_excel=True):
    with open(os.getenv('PATH_DYNAMODB_CHAT_HISTORY'), 'r', encoding='utf-8') as file:
        json_data = json.load(file)

    df_json = pd.DataFrame(json_data)
    df_json['HumanMessage'] = df_json['HumanMessage'].apply(
        lambda x: normalize_unicode_characters(clean_text(x)) if isinstance(x, str) else None)
    df_json['AIMessage'] = df_json['AIMessage'].apply(
        lambda x: normalize_unicode_characters(clean_text(x)) if isinstance(x, str) else None)
    df_json['enrichment_contexts'] = df_json['enrichment'].apply(
        lambda x: [normalize_unicode_characters(context) for context in split_enrichment_contexts(x)] if isinstance(x, str) else [])

    dtype_dict = {'question': str}
    df = pd.read_excel(input_file, dtype=dtype_dict) if use_excel else pd.read_csv(input_file, dtype=dtype_dict)

    df['answer'] = df['answer'].apply(lambda x: normalize_unicode_characters(clean_text(x)))
    df['ground_truths'] = df['ground_truths'].apply(lambda x: normalize_unicode_characters(clean_text(x)))
    df['question'] = df['question'].apply(lambda x: normalize_unicode_characters(clean_text(x)))

    df['question'] = df['question'].astype(str)
    df['answer'] = df['answer'].astype(str)
    df['answer_links'] = df['answer_links'].astype(str)

    matched_results = []
    match_count = 0
    matching_threshold = 0.9

    for index, data_row in tqdm(df.iterrows(), total=df.shape[0], desc="Matching Fortschritt"):
        data_question = data_row['question']
        best_match = None
        highest_similarity = 0

        for _, json_row in df_json.iterrows():
            json_humanmessage = json_row['HumanMessage']
            similarity_score = similar(data_question, json_humanmessage)
            if similarity_score > highest_similarity:
                best_match = json_row
                highest_similarity = similarity_score

        if highest_similarity >= matching_threshold:
            match_count += 1
            data_row['contexts'] = [normalize_unicode_characters(clean_text(context)) for context in best_match['enrichment_contexts']]
            matched_results.append(data_row.to_dict())
        else:
            pass

        tqdm.write(f"Matches gefunden: {match_count}/{index + 1}")

    if match_count == 0:
        logging.warning("Keine Matches gefunden. Script wird beendet.")
        sys.exit(0)

    if use_excel:
        final_excel_path = results_file
        wb = Workbook()
        ws = wb.active
        ws.append(['question', 'ground_truths', 'answer', 'contexts'])

        for record in matched_results:
            ws.append([
                record['question'],
                record['ground_truths'],
                record['answer'],
                '; '.join(record.get('contexts', []))
            ])
        wb.save(final_excel_path)
    else:
        final_csv_path = results_file
        control_file_path = final_csv_path.replace('.csv', '_control.csv')

        with open(final_csv_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            fieldnames = ['question', 'ground_truths', 'answer', 'answer_links', 'contexts']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            writer.writeheader()
            for record in matched_results:
                record = {
                    'question': record['question'],
                    'ground_truths': record['ground_truths'],
                    'answer': record['answer'],
                    'answer_links': record['answer_links'],
                    'contexts': '; '.join(record.get('contexts', []))
                }
                try:
                    writer.writerow(record)
                except Exception as e:
                    logging.error(f"Fehler beim Schreiben des Datensatzes in final_csv_path: {e}\n{record}")
                    continue

        with open(control_file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            fieldnames = ['csv_question', 'human_message', 'matched_answer', 'contexts', 'Übereinstimmungs_Hinweis']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            for record in matched_results:
                human_message = None
                best_similarity_score = 0
                for _, item in df_json.iterrows():
                    current_score = similar(record['question'], item['HumanMessage'])
                    if current_score > best_similarity_score:
                        best_similarity_score = current_score
                        human_message = item['HumanMessage']

                übereinstimmungs_hinweis = "OK"
                if best_similarity_score < 0.7:
                    übereinstimmungs_hinweis = "Check it!"

                try:
                    writer.writerow({
                        'csv_question': record['question'],
                        'human_message': human_message or "Keine Übereinstimmung gefunden",
                        'matched_answer': record['answer'],
                        'contexts': '; '.join(record.get('contexts', [])),
                        'Übereinstimmungs_Hinweis': übereinstimmungs_hinweis
                    })
                except Exception as e:
                    logging.error(f"Fehler beim Schreiben des Datensatzes in control_file_path: {e}\n{record}")
                    continue

# Funktion zur Evaluierung der RAGAS-Ergebnisse
def evaluate_ragas(input_file, use_excel=True):
    try:
        openai_api_key = os.getenv('OPENAI_API_KEY')
        if openai_api_key is None:
            logging.error("OPENAI_API_KEY nicht gefunden.")
            sys.exit(1)
        os.environ["OPENAI_API_KEY"] = openai_api_key

        dtype_dict = {'question': str, 'answer': str}
        df = pd.read_excel(input_file, dtype=dtype_dict) if use_excel else pd.read_csv(input_file, dtype=dtype_dict)
        df = df.applymap(lambda x: normalize_unicode_characters(x) if isinstance(x, str) else x)
        df['question'] = df['question'].astype(str)
        df['answer'] = df['answer'].astype(str)
        df['answer_links'] = df['answer_links'].astype(str)
        df['contexts'] = df['contexts'].apply(lambda x: x if isinstance(x, list) else [x])
        df['ground_truths'] = df['ground_truths'].apply(lambda x: x if isinstance(x, list) else [x])

        valid_rows = []
        for index, row in df.iterrows():
            problem_description = ""
            if not isinstance(row['question'], str):
                problem_description += "Frage ist kein String. "
            if not isinstance(row['answer'], str):
                problem_description += "Antwort ist kein String. "
            if not all(isinstance(item, str) for item in row['contexts']):
                problem_description += "Contexts enthalten Nicht-String-Elemente. "
            if not all(isinstance(item, str) for item in row['ground_truths']):
                problem_description += "Ground Truths enthalten Nicht-String-Elemente. "

            if problem_description:
                logging.warning(f"Zeile {index} übersprungen: {problem_description}")
            else:
                valid_rows.append(row)

        valid_df = pd.DataFrame(valid_rows)
        dataset = Dataset.from_pandas(valid_df)

        result = evaluate(
            dataset,
            metrics=[
                context_relevancy,
                answer_relevancy,
                faithfulness,
                context_recall
            ],
        )
        logging.info("Evaluierung erfolgreich abgeschlossen.")
        result_df = result.to_pandas()
        result_df = result_df.applymap(lambda x: normalize_unicode_characters(x) if isinstance(x, str) else x)
        if use_excel:
            result_df.to_excel(input_file, index=False)
        else:
            result_df.to_csv(input_file, index=False)
    except Exception as e:
        logging.error(f"Fehler bei der Evaluierung: {e}", exc_info=True)
        sys.exit(1)

# ---- START main program ---- #

# Hauptprogramm
use_excel_files = os.getenv('USE_EXCEL') == 'True'
source_file_path = os.getenv('CSV_FILE_PATH_FIXED_DATA')
destination_file_path = os.getenv('PATH_TEST_DATA_RAG_ANSWERS')

# Datei kopieren und umbenennen
copy_and_rename_file(source_file_path, destination_file_path)

# Daten einlesen
data = read_data(destination_file_path, use_excel_files, 'all')

questions, ground_truths = zip(*data)
answers = []
answer_links = []

current_run_user_id = f"USER_ID_RAGAS_{random.randint(1, 9999999)}"
logger.info(f"#################### Running RAGAS script for user ID: {current_run_user_id} ####################\n")
for question in tqdm(questions, desc='Fragen verarbeiten'):
    chat_id, current_run_user_id = create_new_chat(current_run_user_id)
    if chat_id:
        answer, links = get_answer_from_backend(question, chat_id, current_run_user_id)
        answers.append(normalize_unicode_characters(answer) if answer is not None else '')
        answer_links.append(normalize_unicode_characters(links))
    else:
        answers.append('')
        answer_links.append('NA')

# Ergebnisse speichern
if use_excel_files:
    save_results_excel(questions, ground_truths, answers, destination_file_path)
else:
    save_results_csv(questions, ground_truths, answers, destination_file_path, answer_links)

# read from DynamoDB
process_dynamodb_data(current_run_user_id)

# Abgleichen und Anreichern von Daten
process_and_match_data(input_file=os.getenv('PATH_TEST_DATA_RAG_ANSWERS'),
                       results_file=os.getenv('PATH_RESULTS_EVAL_RAGAS'),
                       use_excel=use_excel_files)

# Evaluierung der RAGAS-Ergebnisse
evaluate_ragas(input_file=os.getenv('PATH_RESULTS_EVAL_RAGAS'), use_excel=use_excel_files)

# Funktion zur Konvertierung von CSV zu XLSX
def convert_csv_to_xlsx(CSV_FILE_PATH_FIXED_DATA, xlsx_file_path):
    try:
        logging.info(f'Starting to convert {CSV_FILE_PATH_FIXED_DATA} to XLSX format.')
        csv_data = pd.read_csv(CSV_FILE_PATH_FIXED_DATA, encoding='utf-8')
        csv_data.to_excel(xlsx_file_path, index=False)
        logging.info(f'Successfully converted {CSV_FILE_PATH_FIXED_DATA} to {xlsx_file_path}.')
    except Exception as e:
        logging.error(f'Error converting {CSV_FILE_PATH_FIXED_DATA} to XLSX: {e}')
        sys.exit(1)

final_csv_path = os.getenv('PATH_RESULTS_EVAL_RAGAS')
now = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
final_xlsx_path = final_csv_path.replace('.csv', f'_{now}_fixed_dataset.xlsx')
convert_csv_to_xlsx(final_csv_path, final_xlsx_path)
# ---- END main program ---- #